from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

app = Flask(__name__)
app.secret_key = 'attendify_secret_key_2024'

# Data file paths
DATA_DIR = 'data'
CLASSES_FILE = os.path.join(DATA_DIR, 'classes.json')
ATTENDANCE_FILE = os.path.join(DATA_DIR, 'attendance.json')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')

# Create data directory if it doesn't exist
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

# Teacher credentials
TEACHERS = {
    "teacher1": {"password": "123456", "name": "Dr. Ahmad Hassan"},
    "teacher2": {"password": "123456", "name": "Prof. Sarah Khan"},
    "admin": {"password": "admin123", "name": "Admin User"}
}

def load_data(filename):
    """Load data from JSON file"""
    try:
        with open(filename, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_data(filename, data):
    """Save data to JSON file"""
    with open(filename, 'w') as f:
        json.dump(data, f, indent=2)

@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    username = request.form['username'].strip()
    password = request.form['password'].strip()
    
    if username in TEACHERS and TEACHERS[username]['password'] == password:
        teacher_name = request.form.get('teacher_name', '').strip()
        if not teacher_name:
            teacher_name = TEACHERS[username]['name']
        
        session['user'] = {
            'username': username,
            'name': TEACHERS[username]['name'],
            'display_name': teacher_name
        }
        return redirect(url_for('dashboard'))
    else:
        flash('Invalid username or password', 'error')
        return redirect(url_for('index'))

@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect(url_for('index'))
    
    classes = load_data(CLASSES_FILE)
    return render_template('dashboard.html', 
                         user=session['user'], 
                         classes=classes,
                         current_date=datetime.now().strftime('%Y-%m-%d'))

@app.route('/create_class', methods=['POST'])
def create_class():
    if 'user' not in session:
        return redirect(url_for('index'))
    
    class_name = request.form['class_name'].strip()
    subject = request.form['subject'].strip()
    
    if not class_name or not subject:
        flash('Please fill in all fields', 'error')
        return redirect(url_for('dashboard'))
    
    classes = load_data(CLASSES_FILE)
    
    if class_name in classes:
        flash('Class already exists', 'error')
        return redirect(url_for('dashboard'))
    
    classes[class_name] = {
        'subject': subject,
        'students': []
    }
    
    save_data(CLASSES_FILE, classes)
    flash('Class created successfully!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/add_student', methods=['POST'])
def add_student():
    if 'user' not in session:
        return redirect(url_for('index'))
    
    class_name = request.form['class_name']
    student_name = request.form['student_name'].strip()
    roll_no = request.form['roll_no'].strip()
    
    if not student_name or not roll_no:
        flash('Please fill in all fields', 'error')
        return redirect(url_for('dashboard'))
    
    classes = load_data(CLASSES_FILE)
    
    if class_name not in classes:
        flash('Class not found', 'error')
        return redirect(url_for('dashboard'))
    
    # Check if roll number already exists
    for student in classes[class_name]['students']:
        if student['roll_no'] == roll_no:
            flash('Roll number already exists', 'error')
            return redirect(url_for('dashboard'))
    
    classes[class_name]['students'].append({
        'name': student_name,
        'roll_no': roll_no
    })
    
    save_data(CLASSES_FILE, classes)
    flash('Student added successfully!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/delete_student', methods=['POST'])
def delete_student():
    if 'user' not in session:
        return redirect(url_for('index'))
    
    class_name = request.form['class_name']
    roll_no = request.form['roll_no']
    
    classes = load_data(CLASSES_FILE)
    
    if class_name in classes:
        classes[class_name]['students'] = [
            s for s in classes[class_name]['students'] 
            if s['roll_no'] != roll_no
        ]
        save_data(CLASSES_FILE, classes)
        flash('Student deleted successfully!', 'success')
    
    return redirect(url_for('dashboard'))

@app.route('/delete_class', methods=['POST'])
def delete_class():
    if 'user' not in session:
        return redirect(url_for('index'))
    
    class_name = request.form['class_name']
    classes = load_data(CLASSES_FILE)
    
    if class_name in classes:
        del classes[class_name]
        save_data(CLASSES_FILE, classes)
        flash('Class deleted successfully!', 'success')
    
    return redirect(url_for('dashboard'))

@app.route('/get_students/<class_name>')
def get_students(class_name):
    if 'user' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    classes = load_data(CLASSES_FILE)
    if class_name in classes:
        return jsonify(classes[class_name]['students'])
    return jsonify([])

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    if 'user' not in session:
        return redirect(url_for('index'))
    
    class_name = request.form['class_name']
    date = request.form['date']
    attendance_data = request.form.to_dict()
    
    # Remove non-attendance fields
    attendance_data.pop('class_name', None)
    attendance_data.pop('date', None)
    
    all_attendance = load_data(ATTENDANCE_FILE)
    key = f"{class_name}_{date}"
    all_attendance[key] = attendance_data
    
    save_data(ATTENDANCE_FILE, all_attendance)
    flash('Attendance saved successfully!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/get_attendance/<class_name>/<date>')
def get_attendance(class_name, date):
    if 'user' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    all_attendance = load_data(ATTENDANCE_FILE)
    key = f"{class_name}_{date}"
    return jsonify(all_attendance.get(key, {}))

@app.route('/generate_report', methods=['POST'])
def generate_report():
    if 'user' not in session:
        return redirect(url_for('index'))
    
    class_name = request.form['class_name']
    date = request.form['date']
    
    classes = load_data(CLASSES_FILE)
    all_attendance = load_data(ATTENDANCE_FILE)
    
    if class_name not in classes:
        flash('Class not found', 'error')
        return redirect(url_for('dashboard'))
    
    students = classes[class_name]['students']
    key = f"{class_name}_{date}"
    attendance = all_attendance.get(key, {})
    
    report_data = {
        'class_name': class_name,
        'date': date,
        'students': students,
        'attendance': attendance,
        'present_count': len([v for v in attendance.values() if v == 'present']),
        'absent_count': len([v for v in attendance.values() if v == 'absent']),
        'total_students': len(students)
    }
    
    return render_template('report.html', report=report_data, user=session['user'])

@app.route('/export_excel/<class_name>')
def export_excel(class_name):
    if 'user' not in session:
        return redirect(url_for('index'))
    
    classes = load_data(CLASSES_FILE)
    all_attendance = load_data(ATTENDANCE_FILE)
    
    if class_name not in classes:
        flash('Class not found', 'error')
        return redirect(url_for('dashboard'))
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_name} Attendance"
    
    # Get all dates for this class
    dates = []
    for key in all_attendance.keys():
        if key.startswith(f"{class_name}_"):
            date = key.replace(f"{class_name}_", "")
            dates.append(date)
    
    dates.sort()
    students = classes[class_name]['students']
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create headers
    ws['A1'] = "Roll No"
    ws['B1'] = "Student Name"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    ws['B1'].font = header_font
    ws['B1'].fill = header_fill
    ws['B1'].alignment = center_alignment
    
    # Add date headers
    for i, date in enumerate(dates):
        col = chr(67 + i)  # C, D, E, etc.
        ws[f'{col}1'] = date
        ws[f'{col}1'].font = header_font
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].alignment = center_alignment
    
    # Add summary columns
    summary_col1 = chr(67 + len(dates))
    summary_col2 = chr(68 + len(dates))
    ws[f'{summary_col1}1'] = "Total Present"
    ws[f'{summary_col2}1'] = "Total Absent"
    ws[f'{summary_col1}1'].font = header_font
    ws[f'{summary_col1}1'].fill = header_fill
    ws[f'{summary_col1}1'].alignment = center_alignment
    ws[f'{summary_col2}1'].font = header_font
    ws[f'{summary_col2}1'].fill = header_fill
    ws[f'{summary_col2}1'].alignment = center_alignment
    
    # Add student data
    for row, student in enumerate(students, start=2):
        ws[f'A{row}'] = student['roll_no']
        ws[f'B{row}'] = student['name']
        ws[f'A{row}'].alignment = center_alignment
        
        present_count = 0
        absent_count = 0
        
        # Add attendance for each date
        for i, date in enumerate(dates):
            col = chr(67 + i)
            key = f"{class_name}_{date}"
            attendance = all_attendance.get(key, {})
            status = attendance.get(student['roll_no'], 'N/A')
            
            ws[f'{col}{row}'] = 'P' if status == 'present' else 'A' if status == 'absent' else 'N/A'
            ws[f'{col}{row}'].alignment = center_alignment
            
            # Color coding
            if status == 'present':
                ws[f'{col}{row}'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                present_count += 1
            elif status == 'absent':
                ws[f'{col}{row}'].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                absent_count += 1
        
        # Add summary counts
        ws[f'{summary_col1}{row}'] = present_count
        ws[f'{summary_col2}{row}'] = absent_count
        ws[f'{summary_col1}{row}'].alignment = center_alignment
        ws[f'{summary_col2}{row}'].alignment = center_alignment
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{class_name}_Attendance_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
